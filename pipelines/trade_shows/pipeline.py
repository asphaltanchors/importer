"""
Trade Show Lead Pipeline
Loads trade show lead data from XLSX files with show metadata
"""
import os
import sys
import json
import logging
from pathlib import Path
from typing import Dict, List, Any
from datetime import datetime
import openpyxl
import dlt
from dlt.common.typing import TDataItem

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))
from shared.utils import setup_logging
from shared.database import get_dlt_destination

logger = setup_logging(__name__)


class TradeShowPipeline:
    """Pipeline for loading trade show lead data"""

    def __init__(self, dropbox_path: str = None):
        self.dropbox_path = dropbox_path or os.getenv('DROPBOX_PATH', '/data')
        self.trade_shows_path = Path(self.dropbox_path) / 'trade_shows'

    def discover_shows(self) -> List[Dict[str, Any]]:
        """Discover all trade show directories with leads.xlsx and metadata"""
        shows = []

        if not self.trade_shows_path.exists():
            logger.warning(f"Trade shows directory not found: {self.trade_shows_path}")
            return shows

        for show_dir in self.trade_shows_path.iterdir():
            if not show_dir.is_dir():
                continue

            leads_file = show_dir / 'leads.xlsx'
            metadata_file = show_dir / 'show_metadata.json'

            if not leads_file.exists():
                logger.warning(f"No leads.xlsx found in {show_dir.name}")
                continue

            # Load metadata if available
            metadata = {}
            if metadata_file.exists():
                try:
                    with open(metadata_file, 'r') as f:
                        metadata = json.load(f)
                except Exception as e:
                    logger.error(f"Error loading metadata for {show_dir.name}: {e}")
            else:
                # Create basic metadata from directory name
                metadata = {
                    'show_name': show_dir.name,
                    'show_date': None,
                    'show_location': None,
                    'show_rep': None
                }

            shows.append({
                'directory': show_dir,
                'leads_file': leads_file,
                'metadata': metadata
            })

        logger.info(f"Discovered {len(shows)} trade show(s)")
        return shows

    def load_leads_from_xlsx(self, show_info: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Load leads from XLSX file and enrich with show metadata"""
        leads_file = show_info['leads_file']
        metadata = show_info['metadata']

        logger.info(f"Loading leads from {leads_file}")

        try:
            wb = openpyxl.load_workbook(leads_file, read_only=True)
            ws = wb.active

            # Get headers from first row
            headers = [cell.value for cell in ws[1]]

            leads = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):  # Skip empty rows
                    continue

                # Create lead dict from row
                lead = dict(zip(headers, row))

                # Add show metadata to each lead
                lead['show_name'] = metadata.get('show_name')
                lead['show_date'] = metadata.get('show_date')
                lead['show_location'] = metadata.get('show_location')
                lead['show_rep'] = metadata.get('show_rep')

                # Add load metadata
                lead['load_date'] = datetime.now().isoformat()
                lead['source_file'] = str(leads_file)

                leads.append(lead)

            wb.close()
            logger.info(f"Loaded {len(leads)} leads from {show_info['metadata']['show_name']}")
            return leads

        except Exception as e:
            logger.error(f"Error loading XLSX file {leads_file}: {e}")
            raise

    def normalize_column_names(self, leads: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Normalize column names to match database schema"""
        import hashlib

        column_mapping = {
            'ID': 'lead_id',
            'First name': 'first_name',
            'Last name': 'last_name',
            'Company': 'company',
            'Department': 'department',
            'Job title': 'job_title',
            'Email': 'email',
            'Phone': 'phone',
            'Street address 1': 'address_1',
            'Street address 2': 'address_2',
            'City': 'city',
            'State': 'state',
            'Postal code': 'postal_code',
            'Country': 'country',
            'Source ID': 'source_id',
            'Notes': 'notes',
            'Created': 'created',
            'Updated': 'updated'
        }

        normalized_leads = []
        for lead in leads:
            normalized = {}
            for old_name, new_name in column_mapping.items():
                normalized[new_name] = lead.get(old_name)

            # Generate ID for leads missing one
            if not normalized['lead_id']:
                # Create deterministic ID from lead data
                id_parts = [
                    str(normalized.get('email', '')),
                    str(normalized.get('first_name', '')),
                    str(normalized.get('last_name', '')),
                    str(normalized.get('company', '')),
                    str(lead.get('show_name', ''))
                ]
                id_string = '|'.join(id_parts).lower()
                normalized['lead_id'] = 'gen_' + hashlib.md5(id_string.encode()).hexdigest()[:16]
                logger.info(f"Generated ID for lead: {normalized.get('first_name')} {normalized.get('last_name')} ({normalized.get('email')})")

            # Keep show metadata
            normalized['show_name'] = lead.get('show_name')
            normalized['show_date'] = lead.get('show_date')
            normalized['show_location'] = lead.get('show_location')
            normalized['show_rep'] = lead.get('show_rep')
            normalized['load_date'] = lead.get('load_date')
            normalized['source_file'] = lead.get('source_file')

            normalized_leads.append(normalized)

        return normalized_leads

    def get_leads(self) -> List[Dict[str, Any]]:
        """Get all leads from all trade shows"""
        shows = self.discover_shows()

        if not shows:
            logger.warning("No trade shows found")
            return []

        all_leads = []
        for show_info in shows:
            leads = self.load_leads_from_xlsx(show_info)
            normalized_leads = self.normalize_column_names(leads)
            all_leads.extend(normalized_leads)

        return all_leads


@dlt.resource(
    name="trade_show_leads",
    write_disposition="replace",
    primary_key="lead_id"
)
def trade_show_leads_resource() -> TDataItem:
    """DLT resource for trade show leads"""
    ts_pipeline = TradeShowPipeline()
    leads = ts_pipeline.get_leads()

    for lead in leads:
        yield lead


def run_trade_show_pipeline(mode: str = "replace") -> None:
    """
    Run the trade show pipeline

    Args:
        mode: "replace" to replace all data, "merge" to upsert
    """
    logger.info(f"Starting trade show pipeline in {mode} mode")

    try:
        # Create DLT pipeline
        pipeline = dlt.pipeline(
            pipeline_name="trade_shows_pipeline",
            destination=get_dlt_destination(),
            dataset_name="raw",
        )

        # Run the pipeline
        logger.info("Loading trade show leads...")
        load_info = pipeline.run(
            trade_show_leads_resource(),
            write_disposition=mode
        )

        logger.info(f"Trade show pipeline completed: {load_info}")
        print(f"✅ Trade show pipeline complete: {load_info}")

    except Exception as e:
        logger.error(f"Trade show pipeline failed: {e}")
        print(f"❌ Trade show pipeline failed: {e}")
        raise


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Trade Show Lead Pipeline")
    parser.add_argument(
        "--mode",
        choices=["replace", "merge"],
        default="replace",
        help="Pipeline mode: replace (default) or merge"
    )

    args = parser.parse_args()
    run_trade_show_pipeline(mode=args.mode)
